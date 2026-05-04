"""
Garināšanas modulis — darba uzdevumu ģenerācija
Ievada izejmateriālu un mērķa garumus → aprēķina shēmas → eksports Excel
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import date
from itertools import product

# openpyxl importēts caur pandas to_excel; papildus formatēšanai:
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# KONSTANTES
# ============================================================
KERF_MM = 6.2  # zāģa rezes platums
GARUMA_PIELIKUMS = 10  # +mm pie nominālā garuma
PAKAS_15X60 = 900  # gab/paka standarta izmērs (15 wide × 60 high)


# ============================================================
# SHĒMU APRĒĶINS
# ============================================================
def atrast_shemas(faktiskais_garums_mm, merka_garumi, max_gab_per_dēli=6):
    """
    Atrod visas dzīvotspējīgās griešanas shēmas.

    Atgriež sarakstu ar dict: combo, n_pieces, sum_pieces, kerf_mm,
    aiznem_mm, atlikums_mm, izmantotie_garumi (dict: garums -> skaits).
    """
    shemas = []
    # Kombinācijas: cik gabalus no katra garuma var ņemt vienā dēlī
    # Robežojam max gabalu skaitu, lai nesprāgst
    range_per_length = []
    for g in merka_garumi:
        max_n = min(max_gab_per_dēli, faktiskais_garums_mm // g + 1)
        range_per_length.append(range(int(max_n) + 1))

    seen = set()
    for combo in product(*range_per_length):
        n_total = sum(combo)
        if n_total == 0:
            continue
        if n_total > max_gab_per_dēli:
            continue
        sum_pieces = sum(c * g for c, g in zip(combo, merka_garumi))
        kerf = n_total * KERF_MM
        aiznem = sum_pieces + kerf
        if aiznem > faktiskais_garums_mm:
            continue
        atlikums = faktiskais_garums_mm - aiznem
        # Pieņemam shēmu, ja atlikums < mazākā garuma (citādi var iegūt vēl gabalu)
        if atlikums >= min(merka_garumi):
            continue
        # Combo kā stabils atslēgas
        key = combo
        if key in seen:
            continue
        seen.add(key)
        # Kombinācijas tekstuāls apraksts
        parts = []
        izmantotie = {}
        for c, g in zip(combo, merka_garumi):
            if c > 0:
                parts.append(f"{c}×{g}")
                izmantotie[g] = c
        shemas.append({
            "combo_str": " + ".join(parts),
            "combo": combo,
            "n_pieces": n_total,
            "sum_pieces": sum_pieces,
            "kerf_mm": round(kerf, 1),
            "aiznem_mm": round(aiznem, 1),
            "atlikums_mm": round(atlikums, 1),
            "izmantotie": izmantotie,
        })
    return shemas


def izveleties_shemas(shemas, vajadzigie):
    """
    Vienkārša heuristika — atlasa shēmu kombināciju, kas tuvāk apmierina
    vajadzīgos garumus. Lieto secīgu aizpildi sākot ar lielākajiem garumiem.

    vajadzigie: dict {garums: skaits}
    Atgriež: list of dict {shema, deli_skaits, izveidotais: dict}
    """
    atlikums = dict(vajadzigie)  # ko vēl vajag
    rezultats = []

    # Kārtojam mērķa garumus dilstoši pēc atlikušās vajadzības
    while True:
        # Pārtraucam, ja viss apmierināts
        if all(v <= 0 for v in atlikums.values()):
            break

        # Atrast shēmu, kas dod visvairāk no garuma ar lielāko atlikušo vajadzību
        prioritate = max(atlikums, key=atlikums.get)
        if atlikums[prioritate] <= 0:
            break

        # Filtrējam shēmas, kas izmanto šo garumu
        kandidati = [s for s in shemas if s["izmantotie"].get(prioritate, 0) > 0]
        if not kandidati:
            break

        # Izvēlamies shēmu, kas dod visvairāk no prioritātes garuma
        # un mazāk pārprodukcijas pārējiem
        def skore(s):
            no_prior = s["izmantotie"].get(prioritate, 0)
            # Sods par pārprodukciju
            sods = 0
            for g, n in s["izmantotie"].items():
                if g != prioritate:
                    sods += max(0, n - max(0, atlikums.get(g, 0)) / max(1, no_prior))
            return (no_prior, -sods, -s["atlikums_mm"])

        labakais = max(kandidati, key=skore)
        no_prior_per_dēli = labakais["izmantotie"].get(prioritate, 0)
        # Cik dēļus vajag, lai apmierinātu prioritāti
        deli = -(-atlikums[prioritate] // no_prior_per_dēli)  # ceil division

        # Atjaunina atlikumus
        izveidots = {}
        for g, n in labakais["izmantotie"].items():
            izveidots[g] = n * deli
            atlikums[g] = atlikums.get(g, 0) - n * deli

        rezultats.append({
            "shema": labakais,
            "deli_skaits": deli,
            "izveidots": izveidots,
        })

    return rezultats


# ============================================================
# EXCEL EKSPORTS
# ============================================================
def generet_excel(du_dati):
    """
    Izveido Excel failu no DU datiem.
    du_dati: dict ar visiem ievades un aprēķinu datiem.
    Atgriež BytesIO ar Excel saturu.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Garināšanas uzdevums"

    # Stili
    title_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", start_color="1F4E78")
    section_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", start_color="2E75B6")
    header_font = Font(name="Arial", size=10, bold=True)
    header_fill = PatternFill("solid", start_color="D9E1F2")
    total_font = Font(name="Arial", size=10, bold=True)
    total_fill = PatternFill("solid", start_color="FFF2CC")
    warn_fill = PatternFill("solid", start_color="FFE699")
    label_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    note_font = Font(name="Arial", size=10, italic=True)

    thin = Side(border_style="thin", color="808080")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = "GARINĀŠANAS DARBA UZDEVUMS"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    # Header
    ws["A3"] = "DATUMS"; ws["A3"].font = label_font
    ws["B3"] = du_dati["datums"].strftime("%d.%m.%Y")
    ws["D3"] = "KLIENTS"; ws["D3"].font = label_font
    ws["E3"] = du_dati["klients"]
    ws["A4"] = "TERMIŅŠ"; ws["A4"].font = label_font
    ws["B4"] = du_dati["termins"].strftime("%d.%m.%Y")
    ws["D4"] = "Darba uzdevums NR:"; ws["D4"].font = label_font
    ws["E4"] = du_dati["du_nr"]

    # Tehniskie parametri
    ws.merge_cells("A6:J6")
    ws["A6"] = "TEHNISKIE PARAMETRI"
    ws["A6"].font = section_font; ws["A6"].fill = section_fill; ws["A6"].alignment = left

    params = [
        ("Dēļa nominālais garums", f"{du_dati['nom_garums']} mm"),
        ("Faktiskais garums", f"{du_dati['fakt_garums']} mm"),
        ("Zāģa rezes platums", f"{KERF_MM} mm"),
        ("Materiāls", f"{du_dati['suga']} {du_dati['augstums']} × {du_dati['platums']} mm"),
    ]
    for i, (k, v) in enumerate(params):
        r = 7 + i
        ws.cell(row=r, column=1, value=k).font = label_font
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.cell(row=r, column=4, value=v).alignment = left

    # Izejmateriāls
    ws.merge_cells("A12:J12")
    ws["A12"] = "IZEJMATERIĀLS"
    ws["A12"].font = section_font; ws["A12"].fill = section_fill; ws["A12"].alignment = left

    src_headers = ["Suga", "Augstums", "Platums", "Garums", "Pakas", "Kopā gab", "Kopā m³", "Piegādātājs"]
    for i, h in enumerate(src_headers, start=1):
        c = ws.cell(row=13, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = box

    src_row = [
        du_dati["suga"], du_dati["augstums"], du_dati["platums"], du_dati["nom_garums"],
        du_dati["pakas_skaits"], du_dati["kopa_gab"],
        f"=B14*C14*D14*F14/1000000000", du_dati["piegadatajs"]
    ]
    for i, v in enumerate(src_row, start=1):
        c = ws.cell(row=14, column=i, value=v)
        c.alignment = center; c.border = box
    ws.cell(row=14, column=7).number_format = "0.000"

    # Nepieciešamie garumi
    ws.merge_cells("A16:J16")
    ws["A16"] = "NEPIECIEŠAMIE GARUMI"
    ws["A16"].font = section_font; ws["A16"].fill = section_fill; ws["A16"].alignment = left

    req_headers = ["Nr.", "Augstums", "Platums", "Garums", "Gab", "m³/gab", "Kopā m³"]
    for i, h in enumerate(req_headers, start=1):
        c = ws.cell(row=17, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = box

    for idx, mg in enumerate(du_dati["merka_garumi_dati"]):
        r = 18 + idx
        ws.cell(row=r, column=1, value=idx+1).alignment = center
        ws.cell(row=r, column=2, value=du_dati["augstums"]).alignment = center
        ws.cell(row=r, column=3, value=du_dati["platums"]).alignment = center
        ws.cell(row=r, column=4, value=mg["garums"]).alignment = center
        ws.cell(row=r, column=5, value=mg["gab"]).alignment = center
        ws.cell(row=r, column=6, value=f"=B{r}*C{r}*D{r}/1000000000").number_format = "0.000000"
        ws.cell(row=r, column=6).alignment = center
        ws.cell(row=r, column=7, value=f"=E{r}*F{r}").number_format = "0.0000"
        ws.cell(row=r, column=7).alignment = center
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = box

    n_garumi = len(du_dati["merka_garumi_dati"])
    trow = 18 + n_garumi
    ws.cell(row=trow, column=1, value="KOPĀ").font = total_font
    ws.merge_cells(start_row=trow, start_column=1, end_row=trow, end_column=4)
    for col in range(1, 5):
        ws.cell(row=trow, column=col).fill = total_fill; ws.cell(row=trow, column=col).border = box
    ws.cell(row=trow, column=1).alignment = center
    ws.cell(row=trow, column=5, value=f"=SUM(E18:E{trow-1})").font = total_font
    ws.cell(row=trow, column=5).fill = total_fill; ws.cell(row=trow, column=5).alignment = center
    ws.cell(row=trow, column=5).border = box
    ws.cell(row=trow, column=6).fill = total_fill; ws.cell(row=trow, column=6).border = box
    ws.cell(row=trow, column=7, value=f"=SUM(G18:G{trow-1})").font = total_font
    ws.cell(row=trow, column=7).fill = total_fill; ws.cell(row=trow, column=7).number_format = "0.0000"
    ws.cell(row=trow, column=7).alignment = center; ws.cell(row=trow, column=7).border = box

    # Griešanas shēmas
    sh_start = trow + 2
    ws.merge_cells(f"A{sh_start}:J{sh_start}")
    ws[f"A{sh_start}"] = "GRIEŠANAS SHĒMAS"
    ws[f"A{sh_start}"].font = section_font; ws[f"A{sh_start}"].fill = section_fill; ws[f"A{sh_start}"].alignment = left

    sh_headers = ["Sh.", "Kombinācija", "Gab/dēlī", "Σ pcs", "Rēzes", "Aizņem", "Atlikums", "Dēļi"]
    # Plus per-length output columns
    out_lengths = [mg["garums"] for mg in du_dati["merka_garumi_dati"]]
    sh_headers += [str(l) for l in out_lengths]
    for i, h in enumerate(sh_headers, start=1):
        c = ws.cell(row=sh_start+1, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = box

    for idx, izv in enumerate(du_dati["izveletas_shemas"]):
        r = sh_start + 2 + idx
        s = izv["shema"]
        labels = "ABCDEFGHIJ"[idx] if idx < 10 else str(idx+1)
        ws.cell(row=r, column=1, value=labels).alignment = center
        ws.cell(row=r, column=2, value=s["combo_str"]).alignment = left
        ws.cell(row=r, column=3, value=s["n_pieces"]).alignment = center
        ws.cell(row=r, column=4, value=s["sum_pieces"]).alignment = center
        ws.cell(row=r, column=5, value=s["kerf_mm"]).alignment = center
        ws.cell(row=r, column=6, value=s["aiznem_mm"]).alignment = center
        ws.cell(row=r, column=7, value=s["atlikums_mm"]).alignment = center
        ws.cell(row=r, column=8, value=izv["deli_skaits"]).alignment = center
        # Outputs per length
        for j, l in enumerate(out_lengths):
            col = 9 + j
            n = izv["izveidots"].get(l, 0)
            ws.cell(row=r, column=col, value=n).alignment = center
        for col in range(1, 9 + len(out_lengths)):
            ws.cell(row=r, column=col).border = box

    # Shēmu kopsumma
    n_shemas = len(du_dati["izveletas_shemas"])
    sum_row = sh_start + 2 + n_shemas
    ws.cell(row=sum_row, column=1, value="KOPĀ").font = total_font
    ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=7)
    for col in range(1, 8):
        ws.cell(row=sum_row, column=col).fill = total_fill; ws.cell(row=sum_row, column=col).border = box
    ws.cell(row=sum_row, column=1).alignment = center

    deli_kopa = sum(izv["deli_skaits"] for izv in du_dati["izveletas_shemas"])
    ws.cell(row=sum_row, column=8, value=deli_kopa).font = total_font
    ws.cell(row=sum_row, column=8).fill = total_fill; ws.cell(row=sum_row, column=8).alignment = center
    ws.cell(row=sum_row, column=8).border = box

    for j, l in enumerate(out_lengths):
        col = 9 + j
        kopa = sum(izv["izveidots"].get(l, 0) for izv in du_dati["izveletas_shemas"])
        c = ws.cell(row=sum_row, column=col, value=kopa)
        c.font = total_font; c.fill = total_fill; c.alignment = center; c.border = box

    # Iepakošana
    pack_start = sum_row + 2
    ws.merge_cells(f"A{pack_start}:J{pack_start}")
    ws[f"A{pack_start}"] = "IEPAKOŠANA"
    ws[f"A{pack_start}"].font = section_font; ws[f"A{pack_start}"].fill = section_fill; ws[f"A{pack_start}"].alignment = left

    pack_headers = ["Garums", "Kopā gab", "Pakas izmērs", "Gab/paka", "Pakas", "Pēdējā paka"]
    for i, h in enumerate(pack_headers, start=1):
        c = ws.cell(row=pack_start+1, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = box

    for idx, mg in enumerate(du_dati["merka_garumi_dati"]):
        r = pack_start + 2 + idx
        per_pack = mg.get("per_pack", PAKAS_15X60)
        label = mg.get("pack_label", "15 × 60")
        ws.cell(row=r, column=1, value=mg["garums"]).alignment = center
        ws.cell(row=r, column=2, value=mg["gab"]).alignment = center
        ws.cell(row=r, column=3, value=label).alignment = center
        ws.cell(row=r, column=4, value=per_pack).alignment = center
        ws.cell(row=r, column=5, value=f"=ROUNDUP(B{r}/D{r},0)").alignment = center
        ws.cell(row=r, column=6, value=f"=IF(MOD(B{r},D{r})=0,D{r},MOD(B{r},D{r}))").alignment = center
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = box

    # Piezīmes
    notes_start = pack_start + 2 + len(du_dati["merka_garumi_dati"]) + 1
    ws.merge_cells(f"A{notes_start}:J{notes_start}")
    ws[f"A{notes_start}"] = "PIEZĪMES"
    ws[f"A{notes_start}"].font = section_font; ws[f"A{notes_start}"].fill = section_fill; ws[f"A{notes_start}"].alignment = left

    default_notes = [
        "• Starplikas ik pa 10 rindām, 15 cm no pakas gala",
        "• Var izmantot tos pašus dēļus kā starplikas",
        "• Griezumi perpendikulāri (90°), tolerance ±2 mm",
        "• Katru garumu pakot atsevišķi — nesajaukt!",
        "• Marķēt katru paku: garums, gab, pakas Nr.",
    ]
    if du_dati.get("piezimes"):
        default_notes.append("")
        default_notes.append("Papildus piezīmes:")
        for line in du_dati["piezimes"].split("\n"):
            if line.strip():
                default_notes.append(f"• {line.strip()}")

    for i, n in enumerate(default_notes):
        r = notes_start + 1 + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        ws.cell(row=r, column=1, value=n).alignment = left

    # Footer
    foot = notes_start + 1 + len(default_notes) + 2
    ws.cell(row=foot, column=1, value="Darbus norīko:").font = label_font
    ws.cell(row=foot, column=2, value="_______________________")
    ws.cell(row=foot+1, column=1, value="Operators:").font = label_font
    ws.cell(row=foot+1, column=2, value="_______________________")

    # Kolonu platumi
    widths = {"A":16, "B":24, "C":14, "D":14, "E":12, "F":12, "G":13, "H":11,
              "I":8, "J":8, "K":8, "L":8, "M":8}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Saglabā BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# STREAMLIT UI
# ============================================================
def renderet_garinasanu():
    st.title("📝 Garināšana — darba uzdevumi")
    st.caption("DU ģenerācija ar griešanas shēmām un Excel eksportu")

    # ---- IEVADES FORMA ----
    with st.form("du_forma"):
        st.markdown("### Pamatdati")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            datums = st.date_input("Datums", value=date.today())
        with c2:
            termins = st.date_input("Termiņš", value=date.today())
        with c3:
            klients = st.text_input("Klients", value="Upeslīči")
        with c4:
            piegadatajs = st.text_input("Piegādātājs", value="Toftan")

        c5, c6 = st.columns(2)
        with c5:
            du_nr = st.text_input(
                "Darba uzdevums NR",
                value=f"{datums.strftime('%d%m%y')}-1",
                help="Formāts: DDMMGG-N",
            )
        with c6:
            suga = st.text_input("Suga", value="Skuju koks")

        st.markdown("### Materiāls")
        m1, m2, m3, m4, m5 = st.columns(5)
        with m1:
            augstums = st.number_input("Augstums (mm)", min_value=1, value=14)
        with m2:
            platums = st.number_input("Platums (mm)", min_value=1, value=75)
        with m3:
            nom_garums = st.number_input("Nominālais garums (mm)", min_value=100, value=3000)
        with m4:
            pielikums = st.number_input(
                "Pielikums (+mm)", min_value=0, value=GARUMA_PIELIKUMS,
                help="Cik mm dēlis ir garāks par nominālo"
            )
        with m5:
            pakas_skaits = st.number_input("Pakas (skaits)", min_value=1, value=1)

        kopa_gab = st.number_input(
            "Kopā gab (izejmateriālā)",
            min_value=1, value=11756,
            help="Kopējais dēļu skaits visās pakās"
        )

        st.markdown("### Mērķa garumi un daudzumi")
        st.caption("Ievadi vienā rindā: garums(mm), gab. Vienu rindu uz katru garumu.")
        merka_text = st.text_area(
            "Garumi",
            value="1050, 1660\n940, 1340\n750, 6770\n670, 3240\n540, 17550",
            height=140,
            help="Piem.: 1050, 1660"
        )

        st.markdown("### Iepakošana")
        ip1, ip2 = st.columns(2)
        with ip1:
            std_per_pack = st.number_input(
                "Standarta gab/paka", min_value=1, value=PAKAS_15X60,
                help="15 platums × 60 augstums = 900 gab"
            )
        with ip2:
            isos_dub = st.checkbox(
                "Īsākos garumus pakot 2× kopā", value=True,
                help="540mm un mazākus pakot 2 gab garumā → 1800/paka"
            )
        ip_robeza = st.number_input(
            "Garuma robeža 'īsajiem' (mm)", min_value=0, value=600,
            help="Garumi ≤ šī vērtība tiks pakoti dubultoti"
        )

        piezimes = st.text_area(
            "Papildus piezīmes (neobligāti)",
            placeholder="Piem.: jau izdarīts no 270426-1: 2 pakas × 1800 × 540...",
            height=80,
        )

        sub = st.form_submit_button("🔧 Aprēķināt un sagatavot DU", type="primary", use_container_width=True)

    if not sub:
        st.info("👆 Aizpildi formu un nospied 'Aprēķināt un sagatavot DU'")
        return

    # ---- DATU APSTRĀDE ----
    fakt_garums = nom_garums + pielikums

    # Parsē mērķa garumus
    merka_garumi_dati = []
    for line in merka_text.strip().split("\n"):
        line = line.strip()
        if not line:
            continue
        try:
            parts = [p.strip() for p in line.replace(";", ",").split(",")]
            g = int(parts[0])
            n = int(parts[1])
            per_pack = std_per_pack
            pack_label = "15 × 60"
            if isos_dub and g <= ip_robeza:
                per_pack = std_per_pack * 2
                pack_label = "15 × 60 × 2"
            merka_garumi_dati.append({
                "garums": g, "gab": n,
                "per_pack": per_pack, "pack_label": pack_label,
            })
        except (ValueError, IndexError):
            st.error(f"Nepareiza rinda: '{line}'. Formāts: garums, gab")
            return

    if not merka_garumi_dati:
        st.error("Nav norādīts neviens mērķa garums!")
        return

    # Sakārto pēc dilstoši (lielākie garumi pirmie — vieglāk shēmu meklēšanai)
    merka_garumi_dati.sort(key=lambda x: -x["garums"])
    merka_garumi = [m["garums"] for m in merka_garumi_dati]
    vajadzigie = {m["garums"]: m["gab"] for m in merka_garumi_dati}

    # Atrod un izvēlas shēmas
    visas_shemas = atrast_shemas(fakt_garums, merka_garumi)
    if not visas_shemas:
        st.error("Neizdevās atrast nevienu derīgu griešanas shēmu!")
        return

    izveletas = izveleties_shemas(visas_shemas, vajadzigie)
    if not izveletas:
        st.error("Neizdevās izveidot shēmu plānu!")
        return

    # ---- REZULTĀTU ATTĒLOŠANA ----
    st.success(f"✅ Aprēķins gatavs! Atrastas {len(visas_shemas)} derīgas shēmas, izvēlētas {len(izveletas)}.")

    st.markdown("### 📐 Tehniskie parametri")
    tp1, tp2, tp3, tp4 = st.columns(4)
    tp1.metric("Faktiskais garums", f"{fakt_garums} mm")
    tp2.metric("Zāģa reze", f"{KERF_MM} mm")
    tp3.metric("Materiāls", f"{augstums}×{platums} mm")
    tp4.metric("Pieejami dēļi", f"{kopa_gab:,}".replace(",", " "))

    st.markdown("### 📋 Griešanas shēmas")
    shemas_df = []
    for idx, izv in enumerate(izveletas):
        s = izv["shema"]
        row = {
            "Sh.": "ABCDEFGHIJ"[idx] if idx < 10 else str(idx+1),
            "Kombinācija": s["combo_str"],
            "Aizņem (mm)": s["aiznem_mm"],
            "Atlikums (mm)": s["atlikums_mm"],
            "Dēļi": izv["deli_skaits"],
        }
        for l in merka_garumi:
            row[f"{l}mm"] = izv["izveidots"].get(l, 0)
        shemas_df.append(row)
    st.dataframe(pd.DataFrame(shemas_df), use_container_width=True, hide_index=True)

    # Salīdzinājums ar vajadzīgo
    st.markdown("### 🎯 Iznākums vs. mērķis")
    izn_df = []
    for mg in merka_garumi_dati:
        g = mg["garums"]
        sarazots = sum(izv["izveidots"].get(g, 0) for izv in izveletas)
        starpiba = sarazots - mg["gab"]
        izn_df.append({
            "Garums": f"{g} mm",
            "Mērķis": mg["gab"],
            "Saražots": sarazots,
            "Starpība": f"{starpiba:+d}",
            "Statuss": "✅" if starpiba >= 0 else "⚠️ Trūkst",
        })
    st.dataframe(pd.DataFrame(izn_df), use_container_width=True, hide_index=True)

    # Kopsavilkums
    st.markdown("### 📦 Kopsavilkums")
    deli_izmantoti = sum(izv["deli_skaits"] for izv in izveletas)
    paliek = kopa_gab - deli_izmantoti
    ks1, ks2, ks3 = st.columns(3)
    ks1.metric("Izmantojamie dēļi", f"{deli_izmantoti:,}".replace(",", " "))
    ks2.metric("Pieejami", f"{kopa_gab:,}".replace(",", " "))
    ks3.metric(
        "Paliek malā", f"{paliek:,}".replace(",", " "),
        delta=f"{(paliek*augstums*platums*nom_garums/1e9):.2f} m³",
        delta_color="off",
    )

    if paliek < 0:
        st.error(f"⚠️ Trūkst materiāla: {-paliek} dēļu! Vajag pievienot vairāk pakas.")

    # ---- EXCEL EKSPORTS ----
    st.markdown("### 💾 Eksports")
    du_dati = {
        "datums": datums,
        "termins": termins,
        "klients": klients,
        "piegadatajs": piegadatajs,
        "du_nr": du_nr,
        "suga": suga,
        "augstums": augstums,
        "platums": platums,
        "nom_garums": nom_garums,
        "fakt_garums": fakt_garums,
        "pakas_skaits": pakas_skaits,
        "kopa_gab": kopa_gab,
        "merka_garumi_dati": merka_garumi_dati,
        "izveletas_shemas": izveletas,
        "piezimes": piezimes,
    }

    excel_buf = generet_excel(du_dati)
    fails_nosaukums = f"Garinasanas_DU_{du_nr}.xlsx"

    st.download_button(
        label="📥 Lejupielādēt Excel failu",
        data=excel_buf,
        file_name=fails_nosaukums,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

    # CSV eksports (vienkāršāks)
    csv_buf = BytesIO()
    pd.DataFrame(shemas_df).to_csv(csv_buf, index=False, encoding="utf-8-sig")
    csv_buf.seek(0)
    st.download_button(
        label="📄 Lejupielādēt shēmas kā CSV",
        data=csv_buf,
        file_name=f"Garinasanas_shemas_{du_nr}.csv",
        mime="text/csv",
        use_container_width=True,
    )
